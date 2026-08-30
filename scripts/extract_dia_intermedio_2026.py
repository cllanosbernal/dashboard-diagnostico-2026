#!/usr/bin/env python3
"""Normaliza las planillas DIA Intermedio 2026 sin exponer datos personales."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter
from pathlib import Path
from statistics import mean

import openpyxl


AREA_BY_DIR = {
    "01 LECTURA": "Lectura",
    "02 MATEMATICA": "Matemática",
    "03 CIENCIAS NATURALES": "Ciencias Naturales",
    "04 HISTORIA Y CIENCIAS SOCIALES": "Historia y Cs. Sociales",
}


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFD", value) if unicodedata.category(char) != "Mn"
    )


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("%", "").replace("\u00a0", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def find_row(rows, first_cell: str):
    for index, row in enumerate(rows):
        if row and row[0] == first_cell:
            return index, row
    raise ValueError(f"No se encontró la fila '{first_cell}'")


def find_metadata(rows, label: str):
    for row in rows[:30]:
        if row and row[0] == label:
            return row[1]
    return None


def normalize_course(raw_course: str):
    text = re.sub(r"\s+", " ", str(raw_course or "").strip())
    text = re.sub(r"\s*\(HC-310\)\s*", "", text, flags=re.I).strip()
    match = re.match(r"^(II|I|[1-8])\s*[-°º]?\s*([A-Z])$", text, re.I)
    if not match:
        raise ValueError(f"Curso no reconocido: {raw_course!r}")
    nivel, letra = match.group(1).upper(), match.group(2).upper()
    if nivel in {"I", "II"}:
        label = f"{nivel} {letra}"
        ciclo = "Ens. Media"
        nivel_label = f"{nivel}° Medio"
    else:
        number = int(nivel)
        label = f"{number}°{letra}"
        ciclo = "1° Ciclo" if number <= 4 else "2° Ciclo"
        nivel_label = f"{number}° Básico"
        nivel = str(number)
    return {
        "nivel": nivel,
        "letra": letra,
        "curso": label,
        "ciclo": ciclo,
        "nivel_label": nivel_label,
    }


def course_sort_key(row):
    order = {str(i): i for i in range(1, 9)} | {"I": 9, "II": 10}
    return order.get(row["nivel"], 99), row["letra"], row["area"]


def classify_score(score: float):
    if score < 50:
        return "NL"
    if score < 66.7:
        return "PL"
    return "L"


def normalize_level(value):
    text = strip_accents(str(value or "")).lower().strip()
    match = re.search(r"nivel\s*(i{1,3}|[1-3])", text)
    if not match:
        return None
    token = match.group(1)
    return {"1": "I", "2": "II", "3": "III"}.get(token, token.upper())


def data_rows_from_sheet(path: Path):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    header_index, header = find_row(rows, "Número de Lista")
    return rows, header_index, list(header)


def extract_course_file(path: Path, area: str):
    rows, header_index, header = data_rows_from_sheet(path)
    course = normalize_course(find_metadata(rows, "Curso"))
    last_header = str(header[-1] or "").strip()
    indicator_headers = [str(value).strip() for value in header[2:-1] if value not in (None, "")]
    indicator_values = {name: [] for name in indicator_headers}
    student_totals = []
    official_levels = Counter()
    student_count = 0

    for row in rows[header_index + 1 :]:
        list_number = parse_number(row[0] if row else None)
        student_name = row[1] if len(row) > 1 else None
        if list_number is None or not student_name:
            continue
        student_count += 1
        raw_indicators = list(row[2 : 2 + len(indicator_headers)])
        parsed_indicators = []
        for header_name, value in zip(indicator_headers, raw_indicators):
            if area == "Lectura" and course["nivel"] == "1":
                normalized = strip_accents(str(value or "")).upper().strip()
                parsed = 100.0 if normalized == "L" else 0.0 if normalized == "NL" else None
            else:
                parsed = parse_number(value)
            if parsed is not None:
                parsed = max(0.0, min(100.0, parsed))
                indicator_values[header_name].append(parsed)
                parsed_indicators.append(parsed)

        last_value = row[len(header) - 1] if len(row) >= len(header) else None
        if "PORCENTAJE TOTAL" in strip_accents(last_header).upper():
            total = parse_number(last_value)
        else:
            total = mean(parsed_indicators) if parsed_indicators else None
            level = normalize_level(last_value)
            if level:
                official_levels[level] += 1
        if total is not None:
            student_totals.append(max(0.0, min(100.0, total)))

    indicators = {
        name: round(mean(values), 1) for name, values in indicator_values.items() if values
    }
    if student_totals:
        average = round(mean(student_totals), 1)
    elif indicators:
        average = round(mean(indicators.values()), 1)
    else:
        average = None

    calculated_levels = Counter(classify_score(score) for score in student_totals)
    return {
        "area": area,
        **course,
        "estudiantes": student_count,
        "promedio": average,
        "logro": classify_score(average) if average is not None else None,
        "indicadores": indicators,
        "niveles_oficiales": {key: official_levels.get(key, 0) for key in ("I", "II", "III")},
        "tramos_calculados": {key: calculated_levels.get(key, 0) for key in ("NL", "PL", "L")},
        "fuente": path.name,
    }


def extract_monitoring(source_dir: Path):
    monitoring_dir = source_dir / "00 RESUMEN Y MONITOREO"
    paths = sorted(monitoring_dir.glob("*v2.xlsx")) or sorted(monitoring_dir.glob("*.xlsx"))
    if not paths:
        return [], None
    path = paths[-1]
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    header_index, header = find_row(rows, "Nivel del curso")
    result = []
    for row in rows[header_index + 1 :]:
        if not row or row[0] in (None, "") or row[1] in (None, "") or row[2] in (None, ""):
            continue
        try:
            course = normalize_course(f"{row[0]} {str(row[1]).split()[0]}")
        except ValueError:
            continue
        assigned = int(parse_number(row[6]) or 0)
        completed = int(parse_number(row[7]) or 0)
        result.append(
            {
                **course,
                "area": str(row[2]).strip(),
                "informe_generado": str(row[3]).strip(),
                "accion_pendiente": str(row[4]).strip(),
                "matricula": int(parse_number(row[5]) or 0),
                "asignados": assigned,
                "rindieron": completed,
                "cobertura": round(completed / assigned * 100, 1) if assigned else None,
            }
        )
    return result, path.name


def extract_diagnostic(html_path: Path):
    text = html_path.read_text(encoding="utf-8")
    match = re.search(r"const RAW\s*=\s*(\{.*?\});\s*const TEACHERS", text, flags=re.S)
    if not match:
        raise ValueError("No se pudo extraer RAW diagnóstico desde index.html")
    raw = json.loads(match.group(1))
    by_key = {}
    for area_key, area_name in (("lectura", "Lectura"), ("matematica", "Matemática")):
        for row in raw.get(area_key, []):
            by_key[(area_name, row["label"])] = row
    return by_key


def build_output(source_dir: Path, html_path: Path):
    courses = []
    duplicate_checks = []
    for folder, area in AREA_BY_DIR.items():
        candidates = sorted((source_dir / folder).glob("*.xlsx"))
        primary = [path for path in candidates if path.name.upper().endswith("__PLANILLA.XLSX")]
        extras = [path for path in candidates if path not in primary]
        extracted = [extract_course_file(path, area) for path in primary]
        primary_by_course = {row["curso"]: row for row in extracted}
        for path in extras:
            extra = extract_course_file(path, area)
            primary_row = primary_by_course.get(extra["curso"])
            if primary_row:
                duplicate_checks.append(
                    {
                        "area": area,
                        "curso": extra["curso"],
                        "fuente_primaria": primary_row["fuente"],
                        "fuente_duplicada": extra["fuente"],
                        "coincide_n": primary_row["estudiantes"] == extra["estudiantes"],
                        "coincide_promedio": primary_row["promedio"] == extra["promedio"],
                    }
                )
            else:
                extracted.append(extra)
        courses.extend(extracted)

    courses.sort(key=course_sort_key)
    diagnostic = extract_diagnostic(html_path)
    comparisons = []
    for course in courses:
        if course["area"] not in {"Lectura", "Matemática"}:
            continue
        diag = diagnostic.get((course["area"], course["curso"]))
        diag_average = diag.get("prom") if diag else None
        comparisons.append(
            {
                "area": course["area"],
                "nivel": course["nivel"],
                "letra": course["letra"],
                "curso": course["curso"],
                "ciclo": course["ciclo"],
                "diagnostico": diag_average,
                "intermedio": course["promedio"],
                "variacion": round(course["promedio"] - diag_average, 1)
                if diag_average is not None and course["promedio"] is not None
                else None,
            }
        )

    monitoring, monitoring_file = extract_monitoring(source_dir)
    area_summaries = []
    for area in AREA_BY_DIR.values():
        rows = [course for course in courses if course["area"] == area and course["promedio"] is not None]
        area_summaries.append(
            {
                "area": area,
                "cursos": len(rows),
                "estudiantes_rendidos": sum(row["estudiantes"] for row in rows),
                "promedio_cursos": round(mean(row["promedio"] for row in rows), 1) if rows else None,
                "mejor_curso": max(rows, key=lambda row: row["promedio"])["curso"] if rows else None,
                "mejor_promedio": max(row["promedio"] for row in rows) if rows else None,
                "menor_curso": min(rows, key=lambda row: row["promedio"])["curso"] if rows else None,
                "menor_promedio": min(row["promedio"] for row in rows) if rows else None,
            }
        )

    return {
        "metadata": {
            "establecimiento": "Complejo Educacional Maipú",
            "rbd": 9959,
            "periodo": "DIA Monitoreo Intermedio 2026",
            "fecha_fuentes": "2026-08-28",
            "archivos_xlsx": sum(1 for _ in source_dir.rglob("*.xlsx")),
            "registros_curso_area": len(courses),
            "cursos_unicos": len({row["curso"] for row in courses}),
            "archivo_monitoreo": monitoring_file,
            "metodologia": (
                "Promedio de curso calculado como media de los porcentajes reportados por eje o indicador. "
                "En 1° Básico Lectura se transforma L=100 y NL=0. No se publican nombres de estudiantes."
            ),
        },
        "resumen_areas": area_summaries,
        "cursos": courses,
        "comparaciones": comparisons,
        "aplicacion": monitoring,
        "control_duplicados": duplicate_checks,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", required=True, type=Path)
    parser.add_argument("--html", required=True, type=Path)
    parser.add_argument("--json-output", required=True, type=Path)
    parser.add_argument("--js-output", required=True, type=Path)
    args = parser.parse_args()

    result = build_output(args.source_dir, args.html)
    args.json_output.parent.mkdir(parents=True, exist_ok=True)
    args.js_output.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(result, ensure_ascii=False, separators=(",", ":"))
    args.json_output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.js_output.write_text(f"const DIA_INTERMEDIO_2026 = {payload};\n", encoding="utf-8")
    print(json.dumps(result["metadata"], ensure_ascii=False, indent=2))
    print(json.dumps(result["resumen_areas"], ensure_ascii=False, indent=2))
    print(json.dumps(result["control_duplicados"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
